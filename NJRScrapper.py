import os
import math
import winsound
import openpyxl
import PyPDF2
import shutil
import shelve
import datetime
import traceback
from datetime import date
from datetime import timedelta
import geopandas
import logging
import zipfile
from send2trash import send2trash
from matplotlib import pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.ticker as ticker
import matplotlib.dates as mdates
import requests
import re
import time
import pandas as pd
import plotly.express as px
from tabulate import tabulate
from bs4 import BeautifulSoup
from twilio.rest import Client
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementNotVisibleException
from selenium.common.exceptions import NoSuchElementException


class Scraper:

    state_dict = ''
    current_data = ''
    no_of_runs = 1
    event_log = {0: {'Run Type': 'NJR10k', 'Latest Available Data': 'June 2023', 'Run Time': 'N.A',
                     'Run Date': '2023-08-20', 'Days Between Update': 0}}

    def __init__(self):
        # Change the directory to store the temporary Selenium files to be processed
        os.chdir('C:\\Users\\Omar\\Desktop\\Python Temp Folder')
        Scraper.state_dictionary()
        Scraper.create_event_log()
        # Empty list to store the found cities and counties
        self.__counties = []
        self.__towns = []
        # The years and months for the real estate data I'm looking for
        self.__years = ['2019', '2020', '2021', '2022', '2023']
        self.njrdata()
        self.__months = {'01': 'January', '02': 'February',
                         '03': 'March', '04': 'April',
                         '05': 'May', '06': 'June',
                         '07': 'July', '08': 'August',
                         '09': 'September', '10': 'October',
                         '11': 'November', '12': 'December'
                         }

    """ 
    ______________________________________________________________________________________________________________
                                    Use this section to house the decorator functions
    ______________________________________________________________________________________________________________
    """

    @staticmethod
    def logger_decorator(original_function):
        def wrapper(*args, **kwargs):
            logger = logging.getLogger(original_function.__name__)
            logger.setLevel(logging.DEBUG)
            logger.propagate = False
            # Create the FileHandler() and StreamHandler() loggers
            f_handler = logging.FileHandler(
                original_function.__name__ + ' ' + str(datetime.datetime.today().date()) + '.log')
            f_handler.setLevel(logging.DEBUG)
            c_handler = logging.StreamHandler()
            c_handler.setLevel(logging.INFO)
            # Create formatting for the loggers
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                                          datefmt='%d-%b-%y %H:%M:%S')
            # Set the formatter for each handler
            f_handler.setFormatter(formatter)
            c_handler.setFormatter(formatter)
            logger.addHandler(f_handler)
            logger.addHandler(c_handler)

            kwargs['logger'] = logger
            kwargs['f_handler'] = f_handler
            kwargs['c_handler'] = c_handler

            result = original_function(*args, **kwargs)

            if result is None:
                pass
            else:
                return result
        return wrapper

    @staticmethod
    def run_main(original_function):
        def wrapper(*args, **kwargs):

            # Formulate all the date variables
            todays_date = datetime.datetime.today().date()
            data_avail = Scraper.current_data
            temp_date = str(todays_date).split('-')
            day = int(temp_date[2])
            month = int(temp_date[1])
            year = temp_date[0]
            current_run_date = datetime.datetime.strptime(year + '-' + temp_date[1] + '-' + '24', "%Y-%m-%d").date()

            # Logic for calculating the next date to run main()
            if day < 24:
                next_run_date = year + '-' + temp_date[1] + '-' + '24'
            elif day >= 24:
                if data_avail == Scraper.event_log[obj.no_of_runs - 1]['Latest Available Data']:
                    next_run_date = year + '-' + temp_date[1] + '-' + '24'
                else:
                    if month in [1, 2, 3, 4, 5, 6, 7, 8]:
                        nm = str(month + 1)
                        next_month = '0' + nm
                        next_run_date = year + '-' + next_month + '-' + '24'
                    elif month in [9, 10, 11]:
                        next_month = str(month + 1)
                        next_run_date = year + '-' + next_month + '-' + '24'
                    elif month == 12:
                        next_month = '01'
                        year = str(int(temp_date[0]) + 1)
                        next_run_date = year + '-' + next_month + '-' + '24'

            next_run_date = datetime.datetime.strptime(next_run_date, "%Y-%m-%d").date()
            if todays_date >= current_run_date:
                if data_avail == Scraper.event_log[Scraper.no_of_runs - 1]['Latest Available Data']:
                    sleep_time = timedelta(days=1)
                    Scraper.waiting(sleep_time)

                    return 'RESTART'

                else:
                    good_to_go = original_function(*args, **kwargs)

                return good_to_go

            elif current_run_date < todays_date < next_run_date:
                if todays_date < next_run_date:
                    sleep_time = next_run_date - todays_date
                    Scraper.waiting(sleep_time)

                    return 'RESTART'

        return wrapper

    @staticmethod
    def run_quarterly_statistics(todaysdate):
        # I need to create some local variables here to keep runtimes in check
        quarterly_dict = {
            'Q1': {'Run Date': 'April 24'},
            'Q2': {'Run Date': 'July 24'},
            'Q3': {'Run Date': 'October 24'},
            'Q4': {'Run Date': 'January 24'}
        }

        def wrapper(original_function):
            def internal_wrapper(*args, **kwargs):

                current_year = datetime.datetime.today().year

                for k,v in quarterly_dict.items():

                    run_date = datetime.datetime.strptime(quarterly_dict[k]['Run Date'] + ' ' + str(current_year),
                                                          "%B %d %Y")
                    if todaysdate == run_date:
                        results = original_function(*args, **kwargs)
                    elif date != run_date:
                        continue
                    else:
                        for k1,v1 in quarterly_dict.items():
                            if todaysdate > datetime.datetime.strptime(quarterly_dict[k1]['Run Date'] + ' '
                                                                       + str(current_year), "%B %d %Y"):
                                continue
                            else:
                                results = f"The next quarterly statistics run will be for {k1} on {quarterly_dict[k1]['Run Date']}"

                return results
            return internal_wrapper
        return wrapper


    """ 
    ______________________________________________________________________________________________________________
                            Use this section to house the instance, class and static functions
    ______________________________________________________________________________________________________________
    """

    # Function which scrapes the cities and counties from the njrealtor 10k state page
    def area_results(self, soup):
        """
        Function which accepts a BeautifulSoup object to then parse and find the cities
        and counties located in New Jersey
        :param soup: BeautifulSoup object
        :return: None
        """
        area = soup.find('select', id="lmuArea").children
        for obj1 in area:
            newobj = obj1.get_text()
            if newobj in ['Select an area...', 'Entire State', '\n']:
                continue
            else:
                if 'County' in newobj:
                    # There are multiple towns in different counties with the same name. Their county is attached
                    # and need to be separated from the target values
                    if '/' in newobj:
                        # Do not split the city name if it has '/' in it. It signifies that there is more than 1 city
                        # with that name and shows the county it belongs to
                        # newobj = newobj.split('/')
                        self.__towns.append(newobj)
                    else:
                        self.__counties.append(newobj)
                else:
                    self.__towns.append(newobj)

    def check_county(self, pdf_text, town):
        if '(' in pdf_text:
            county = pdf_text.split('(')[0].strip()
            if county in self.__counties:
                real_county = county
        elif '(' not in pdf_text:
            county = pdf_text.strip()
            if county in self.__counties:
                real_county = county
            else:
                real_county = Scraper.find_county(town)

        return real_county

    @classmethod
    def check_results(cls):
        """
        A classmethod which checks the event log for the latest available data and current available
        data on the NJ Realtor website.
        The "start" variable will house the most recent month of data scraped by the program
        The "finish" variable will house last known month for data that needs to be scraped

        Returns the start and finish variables as lists
        """

        event_log_list = list(cls.event_log.keys())
        last_run_num = event_log_list[-1]
        last_data_avail = cls.event_log[last_run_num]['Latest Available Data']

        start = last_data_avail.split()
        finish = cls.current_data.split()

        return start, finish

    @staticmethod
    def cloropleth_maps_state(filename):
        filename = filename
        df = pd.read_excel(filename, sheet_name='All Months')
        years = df['Year'].unique().tolist()
        temp_df = df.copy().sort_values(by='Dates')
        target_df = temp_df[(temp_df['Year'] == years[-1]) & (temp_df['Quarter'] == 'Q3') & (temp_df['Month'] == 'August')]
        target_list = ['Closed Sales', 'Closed Sales YoY', 'Monthly Inventory', 'Monthly Inventory YoY',
                       'Median Sales Price', 'Median Sales Price YoY']

        new_jersey = geopandas.read_file('F:\\Real Estate Investing\\JQH Holding Company LLC\\Python Coding for Real Estate\\Municipal_Boundaries_of_NJ (2).geojson')

        for target in target_list:
            fig = px.choropleth(target_df, geojson=new_jersey, locations="City", color=target,
                                featureidkey='properties.MUN', projection='mercator',
                                hover_name='City')
            # hover_data=[fill this in later], color_discrete_map='Choose color later'
            fig.update_geos(fitbounds="locations", visible=False)
            # fig.update_layout(margin={"r": 0, "t": 0, "l": 0, "b": 0})
            fig.show()

    # def cloropleth_maps_counties(self, filename):
    #     filename = filename
    #     df = pd.read_excel(filename)
    #     county_list = df['County'].unique().tolist()
    #     df.groupby('Counties')
    #
    #     target_list = ['Closed Sales', 'Closed Sales YoY', 'Monthly Inventory', 'Monthly Inventory YoY',
    #                    'Median Sales Price', 'Median Sales Price YoY']
    #
    #     new_jersey = geojson.read_file(geojson_municipal_filename)

        # for county in county_list:
        #     for target in target_list:
        #         fig = px.choropleth(df[df['County' == county]], geojson=new_jersey, locations="County", color=target,
        #                             featureidkey='properties.mun_label', projection='mercator',
        #                             hover_name='City', hover_data=[fill this in later],
        #                             color_discrete_map='Choose color later')
        #         fig.update_geos(fitbounds="locations", visible=False)
        #         fig.update_layout(margin={"r": 0, "t": 0, "l": 0, "b": 0})
        #         fig.show()

    @logger_decorator
    def corrupted_files(self, corrupt_list, **kwargs):
        """
        Function which captures possibly corrupted pdf files and re-downloads them
        :param corrupt_list: List of pdf names which were possiblly corrupted during the njr10k/update_njr10k download
        :param kwargs: keyword args which hold the logger function variables
        :return: possble_corrupted_files: List of corrupted files which couldn't be downloaded
        """

        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']
        corrupt_dict = {}
        possible_corrupted_files = []

        # Extract the town, month and year from the file name and put it in a dictionary for further processing
        # Real estate data pulled for the year 2019 seems to not be available for some areas so we'll skip these files.
        try:
            corrupt_list = corrupt_list
            assert len(corrupt_list) > 0, 'There are no corrupted files. The Extract RE Data function will now initiate...'
        except AssertionError as AE:
            logger.info(f'{AE}')
            return ['No Corrupted Files']
        else:
            for n, i in enumerate(corrupt_list):
                info = i.rstrip('.pdf').split(' ')
                town = info[0:len(info) - 2]
                if len(town) > 1:
                    if 'County' in town:
                        # This means the city name is a duplicate and needs to have the county distinguished
                        # For example: ['Franklin', 'Twp', 'Gloucester', 'County']
                        # --------> ['Franklin', 'Twp', '/', 'Gloucester', 'County']
                        town.insert(town.index('County') - 1, '/')
                        town = ' '.join(town)
                    else:
                        town = ' '.join(town)
                else:
                    town = info[0]

                month = info[-2]
                year = info[-1]

                if year == '2019':
                    # Skip all corrupted files from 2019. That data is not available
                    possible_corrupted_files.append(i)
                    continue
                else:
                    for idx, i1 in enumerate(self.__towns):
                        if town in i1:
                            town = self.__towns[idx]
                            corrupt_dict[n] = [town, month, year]

            # UnboundLocalError produced without this checkpoint.
            if len(corrupt_dict) < 1:
                if len(possible_corrupted_files) > 0:
                    logger.info('All corrupted files have been captured. The Extract RE Data function will now initiate...')
                    time.sleep(0.5)
                    return possible_corrupted_files
                else:
                    logger.info('There are no corrupted files. The Extract RE Data function will now initiate...')
                    time.sleep(0.5)
                    return ['No Corrupted Files']

            base_url = 'http://njar.stats.10kresearch.com/docs/lmu/'

            with requests.Session() as session:
                username, pw = Scraper.get_us_pw('NJRealtor')

                payload1 = {'rd': '10',
                            'passedURL': '/goto/10k/',
                            'case': '',
                            'LoginEmail': username,
                            'LoginPassword': pw,
                            'LoginButton': 'Login'}

                params = {'src': 'Page'}
                # months = list(self.__months.keys())

                # months_tracker = manager.counter(total=len(months), desc=f'Year:', unit='Months')

                url = 'https://www.njrealtor.com/login/?rd=10&passedURL=/goto/10k/'
                url2 = 'https://www.njrealtor.com/ramco-api/web-services/login_POST.php'

                session.get(url)  # Requesting initial log-in page
                session.post(url2, data=payload1)  # response object for initially logging into website

                try:
                    for k, v in corrupt_dict.items():

                        # Redundant checker if a 2019 file has slipped through the first check
                        if v[2] == '2019':
                            continue
                        else:
                            y = v[2]
                        for k1, v1 in self.__months.items():
                            if v[1] in v1:
                                m = k1

                        url3, new_filename = self.create_url_and_pdfname(base_url, y, m, v[0])

                        Scraper.download_pdf(session, url3, params, new_filename, possible_corrupted_files,
                                             logger)

                except IOError:
                    """An OS Error has occurred """
                    logger.exception(f'IOError has Occurred')

                except requests.exceptions.HTTPError as rht:
                    """An HTTP error occurred."""
                    logger.exception(f'An HTTP has Occurred: {rht}')

                except requests.exceptions.Timeout as ret:
                    """The request timed out.
                    Catching this error will catch both
                    :exc:`~requests.exceptions.ConnectTimeout` and
                    :exc:`~requests.exceptions.ReadTimeout` errors.
                    """
                    logger.exception(f'The Request Has Timed Out: {ret}')

                except requests.exceptions.InvalidURL as rei:
                    """The URL provided was somehow invalid."""
                    logger.exception(f'The URL Provided Was Invalid: {rei}')

                except requests.exceptions.RetryError as rer:
                    """Custom retries logic failed"""
                    logger.exception(f'Custom Retries Logic Failed: {rer}')

                except requests.exceptions.StreamConsumedError as res:
                    """The content for this response was already consumed."""
                    logger.exception(f'The Content For This Response Was Already Consumed: {res}')

                except requests.exceptions.ContentDecodingError as rec:
                    """Failed to decode response content."""
                    logger.exception(f'Failed to Decode Response Content: {rec}')

                except requests.exceptions.ChunkedEncodingError as rece:
                    """The server declared chunked encoding but sent an invalid chunk."""
                    logger.exception(f'Invalid Chunk Encoding: {rece}')

                except Exception as e:
                    logger.exception(f'An Error Has Unhandled Occurred: {e} ')

                else:

                    logger.removeHandler(f_handler)
                    logger.removeHandler(c_handler)
                    logging.shutdown()
                    logger.info('All corrupted files have been captured. The Extract RE Data function will now initiate...')
                    time.sleep(0.5)

            return possible_corrupted_files

    @classmethod
    def create_event_log(cls):
        """
        Classmethod which is run during class initialization update the class variable "event log" with
        the class' run history and updates the class variable "no_of_runs". In the event there isn't a shelf
        file available with event log history, a new event log dictionary is created.
        :return: None
        """
        save_path = 'F:\\Python 2.0\\Projects\\Real Life Projects\\NJR Scrapper\\Saved Data\\NJ Scrapper Data Dictionary_v2.dat'
        if os.path.exists(save_path):
            os.chdir('F:\\Python 2.0\\Projects\\Real Life Projects\\NJR Scrapper\\Saved Data')
            try:
                with shelve.open('NJ Scrapper Data Dictionary_v2', writeback=True) as saved_data_file:
                    if saved_data_file['Event Log']:
                        cls.event_log = saved_data_file['Event Log']
                        runs_list = [i for i in cls.event_log.keys()]
                        Scraper.duplicate_eventlog_check()
                        cls.no_of_runs = runs_list[-1] + 1

                os.chdir('C:\\Users\\Omar\\Desktop\\Python Temp Folder')

            except KeyError:
                os.chdir('C:\\Users\\Omar\\Desktop\\Python Temp Folder')
                key_names = ['Run Type', 'Latest Available Data', 'Run Time', 'Run Date', 'Days Between Update']
                if cls.event_log == {}:
                    cls.event_log.setdefault(cls.no_of_runs, {})

                    for kn in key_names:
                        cls.event_log[cls.no_of_runs].setdefault(kn, '')

    def create_url_and_pdfname(self, base_url, year_var, month_var, town_var):

        city_list = town_var.split(' ')
        merged_city_name = ''.join(city_list)

        if '/' not in city_list:

            new_url = base_url + year_var + '-' + month_var + '/x/' + merged_city_name
            new_filename = " ".join([' '.join(city_list), self.__months[month_var], year_var]) + ".pdf"

        elif '/' in city_list:
            merged_city_name = '%2F'.join(merged_city_name.split('/'))
            del city_list[city_list.index('/')]
            new_url = base_url + year_var + '-' + month_var + '/x/' + merged_city_name
            new_filename = " ".join([' '.join(city_list), self.__months[month_var], year_var]) + ".pdf"

        return new_url, new_filename

    @staticmethod
    @logger_decorator
    def CreateZip(*args, **kwargs):
        """
        Staticmethod which creates a zipfile of all previously downloaded and sorted real estate pdfs
        :param args: None
        :param kwargs: Keyword argument dictionary which houses the logger function variables
        :return: zipname
        """
        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']
        os.chdir('C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files')
        zip_folder = 'F:\\Python 2.0\\Projects\\Real Life Projects\\NJR Scrapper\\NJR10k Zips'
        folder = 'C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files'
        zipname = 'NJR10k ' + str(datetime.datetime.now()) + '.zip'
        newZip = zipfile.Zipfile(zipname, 'w')

        # filenames = os.listdir(folder)
        for root, dirs, filenames in os.walk(folder):
            if len(filenames) > 0:
                # Insert logger here stating 'Zipping {Year_Number} Folder'

                newZip.write(root)
                for filename in filenames:
                    # # Insert logger here stating 'Zipping {Filename}'
                    newZip.write(filename)
                    logger.info(f'{filename} has been sent to the zipped folder')

        newZip.close()
        # Move created zipfile to storage folder
        shutil.move(newZip, zip_folder)
        # Check if the PDF Temp Files folder in the Python Temp Folder still exists. If not create it again
        if os.path.exists(folder):
            pass
        else:
            os.makedirs('C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files')

        logger.removeHandler(f_handler)
        logger.removeHandler(c_handler)
        logging.shutdown()

        return zipname

    # Function which scrapes the current month of data available
    def current_data_avail(self, soup):
        """
        Function which accepts a BeautifulSoup object and scrapes the most recent data available to download
        and assigns the value to the current_data class variable
        :param soup: BeautifulSoup object
        :return: None
        """
        current_results = soup.find('select', id="lmuTime").children
        current_results = list(current_results)
        month_year = current_results[2].get_text()
        target = month_year.split(' ')
        year = target[1]
        # month = target[0]
        if year not in self.__years:
            self.__years.append(year)
            main_dictionary[year] = {}
            full_year[year] = {}

        Scraper.current_data = month_year

    # Creates the initial dictionary the scraped city data will be stored
    @staticmethod
    def data_na(town, month, year, main_dict, year_dict):
        """
        Staticmethod which assigns default values of 0, 0.0 and N/A to variables used in the extraction function
        for real estate pdfs which were found to have corrupted data
        :param town: str variable of the name of the town
        :param month: str variable of the month of the target data
        :param year: str variable of the year of the target data
        :param main_dict:
        :param year_dict:
        :return: None
        """

        current_year = year
        city = town
        county = 'N.A'
        quarter = Scraper.quarter(month)
        new_listings_current = 0
        new_listings_fy = 0
        new_listings_per_change = 0.0
        new_listings_per_change_fy = 0.0
        closed_sales_current = 0
        closed_sales_fy = 0
        closed_sales_per_change = 0.0
        closed_sales_per_change_fy = 0.0
        dom_current = 0
        dom_fy = 0
        dom_per_change = 0.0
        dom_per_change_fy = 0.0
        median_sales_fy = 0
        median_sales_current = 0
        median_sales_per_change = 0.0
        median_sales_per_change_fy = 0.0
        percent_lpr_current = 0.0
        percent_lpr_fy = 0.0
        percent_lpr_per_change = 0.0
        percent_lpr_per_change_fy = 0.0
        inventory_current = 0
        inventory_fy = 0
        inventory_per_change = 0.0
        inventory_per_change_fy = 0.0
        supply_current = 0.0
        supply_fy = 0
        supply_per_change = 0.0
        supply_per_change_fy = 0.0

        category_list = ['City', 'County', 'Quarter', 'Month', 'Year', 'New Listings',
                         'New Listing % Change (YoY)', 'Closed Sales',
                         'Closed Sale % Change (YoY)', 'Days on Markets', 'Days on Market % Change (YoY)',
                         'Median Sales Prices',
                         'Median Sales Price % Change (YoY)', 'Percent of Listing Price Received',
                         'Percent of Listing Price Receive % Change (YoY)', 'Inventory of Homes for Sales',
                         'Inventory of Homes for Sale % Change (YoY)', 'Months of Supply',
                         'Months of Supplies % Change (YoY)']

        variable_list = [city, county, quarter, month, current_year, new_listings_current, new_listings_per_change,
                         closed_sales_current,
                         closed_sales_per_change, dom_current, dom_per_change, median_sales_current,
                         median_sales_per_change, percent_lpr_current, percent_lpr_per_change,
                         inventory_current, inventory_per_change, supply_current, supply_per_change]

        fy_variable_list = [city, county, month, current_year, new_listings_fy, new_listings_per_change_fy,
                            closed_sales_fy, closed_sales_per_change_fy, dom_fy, dom_per_change_fy, median_sales_fy,
                            median_sales_per_change_fy, percent_lpr_fy, percent_lpr_per_change_fy,
                            inventory_fy, inventory_per_change_fy, supply_fy, supply_per_change_fy]

        if main_dict[current_year] == {}:
            for idx, n in enumerate(category_list):
                main_dict[current_year].setdefault(n, [])
                main_dict[current_year][n].append(variable_list[idx])
        else:
            for idx, n in enumerate(category_list):
                main_dict[current_year][n].append(variable_list[idx])
        if month == 'December':
            category_list1 = category_list[:]
            del category_list1[2]

            if year_dict[current_year] == {}:
                for idx, n in enumerate(category_list1):
                    year_dict[current_year].setdefault(n, [])
                    year_dict[current_year][n].append(fy_variable_list[idx])
            else:
                for idx, n in enumerate(category_list1):
                    year_dict[current_year][n].append(fy_variable_list[idx])

    # Function which calculates the difference between the current download date and previous date
    # Use this to calculate the average amount of time it takes between new update periods
    @classmethod
    def daysuntilupdate(cls):
        """
        Classmethod which returns a timedelta object that depicts the amount of days between
        the program's last run and current update
        :return: delta (timedelta object)
        """
        current = cls.no_of_runs
        previous = current - 1
        current_date = datetime.datetime.now()
        previous_date = datetime.datetime.strptime(cls.event_log[previous]['Run Date'], "%a %b %d %H:%M:%S %Y")
        delta = current_date - previous_date

        return delta.days

    @staticmethod
    def download_pdf(session_var, pdf_url, params_dict, pdf_name, corrupted_files_list, logger):

        with session_var.get(pdf_url, params=params_dict, stream=True) as reader, open(pdf_name, 'wb') as writer:
            for chunk in reader.iter_content(chunk_size=1000000):
                # Casting the bytes into a str type
                # and slicing the first 20 characters to check if 'PDF' is in
                check_pdf = str(chunk)[:20]
                # print(check_pdf)
                if 'PDF' in check_pdf:
                    writer.write(chunk)
                else:
                    logger.warning(f'WARNING! {pdf_name} is possibly a corrupted file')
                    corrupted_files_list.append(pdf_name)

    # Functon which checks if there's a duplicate vector for the current input year
    # Don't need a try-except block because any errors will be caught by the outter method
    @staticmethod
    def duplicate_vector_check(pdfname, current_vector, year, main_dict):
        """
        Staticmethod which checks the current data vector against each row of the real estate dictionary
        to be sure that no duplicate data is stored in the dictionary
        :param pdfname: Name of the target pdf
        :param current_vector: latest vector of data added to the real estate dictionary
        :param year: key value 'year' used to access the nested year dictionaries
        :param main_dict
        :return: None
        """
        i = len(main_dict[year]['City'])
        number = i - 1
        if number > 0:
            for r in range(number, 0, -1):
                previous_vector = []
                for k in main_dict[year].keys():
                    previous_vector.append(main_dict[year][k][r])

                assert current_vector != previous_vector, f'Duplicate Error: {pdfname} & {previous_vector[0]} ' \
                                                          f'{previous_vector[3]} {previous_vector[4]}.pdf'
        else:
            previous_vector = []
            for k in main_dict[year].keys():
                previous_vector.append(main_dict[year][k][number])

            assert current_vector != previous_vector, f'Duplicate Error: {pdfname} & {previous_vector[0]} ' \
                                                      f'{previous_vector[3]} {previous_vector[4]}.pdf'

    @classmethod
    def duplicate_eventlog_check(cls):
        """
        Classmethod which is run during the class initialization in the create_event_log function. Function checks the
        event log to make sure there are no duplicate logs stored in the dictionary
        :return: None
        """
        i = len(cls.event_log)
        number = i - 1
        if number > 0:
            for r in range(number, 0, -1):
                current_vector_data = [cls.event_log[r]['Run Type'], cls.event_log[r]['Latest Available Data']]
                previous_vector_data = [cls.event_log[r-1]['Run Type'], cls.event_log[r-1]['Latest Available Data']]
                try:
                    assert current_vector_data != previous_vector_data
                except AssertionError:
                    del cls.event_log[r]
                    cls.no_of_runs -= 1
                else:
                    pass

    # This is an instance method because I'm using a static method inside the function which may not be able
    def event_log_update(self, name, run_time, logger):
        """
        Instance method which updates the event log with runtime data of the most recent NJR10k download.
        Stores the type of downlaod/update which was run, the length of the download runtime, current date and
        length in time between the previous and current program runs
        :param name: Name of the function ran (njr10k or update_njr10k)
        :param run_time: time object which value is the download function's run time
        :param logger: logger function which will return event log to ther logger file
        :return: None
        """
        Scraper.event_log.setdefault(Scraper.no_of_runs, {'Run Type': '', 'Latest Available Data': '',
                                                          'Run Time': '', 'Run Date': '', 'Days Between Update': ''})
        Scraper.event_log[Scraper.no_of_runs]['Run Type'] = name
        Scraper.event_log[Scraper.no_of_runs]['Latest Available Data'] = Scraper.current_data
        Scraper.event_log[Scraper.no_of_runs]['Run Time'] = str(run_time)
        Scraper.event_log[Scraper.no_of_runs]['Run Date'] = time.ctime()
        Scraper.event_log[Scraper.no_of_runs]['Days Between Update'] = Scraper.daysuntilupdate()

        logger.info(f'New Event Log Created:\n'
                    f"\nRun Date: {Scraper.event_log[Scraper.no_of_runs]['Run Date']}"
                    f"\nRun #: {Scraper.no_of_runs}"
                    f"\nRun Type: {Scraper.event_log[Scraper.no_of_runs]['Run Type']}"
                    f"\nLatest Available Data: {Scraper.event_log[Scraper.no_of_runs]['Latest Available Data']}"
                    f"\nRun Time: {Scraper.event_log[Scraper.no_of_runs]['Run Time']}"
                    f"\nDays Between Update: {Scraper.event_log[Scraper.no_of_runs]['Days Between Update']}")

    # Function which extracts the month, current and previous year, new listing, closing sales, DOM, median sales, etc
    # Data will then be stored in a dictionary
    @logger_decorator
    def extract_re_data(self, pdfname, possible_corrupted_list, main_dict, year_dict, update=None, **kwargs):
        """
        Function which reads the pdfname name arg and extracts the real estate data from that pdf and stores
        it in the global main_dictionary variable
        :param pdfname: Name of the target pdf
        :param possible_corrupted_list: list variable which stores the name of possibly corrupted files
        :param update: Allows for dynamic directory changing if argument is equal to 'Yes'
        :param main_dict:
        :param year_dict:
        :param kwargs: Keyword argument dictionary which houses the logger function variables
        :return: None
        """

        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']

        # For municipalities with the same name and located in multiple counties,
        # the county name will be unpacked in the town var and needs to be extracted to use as an additional
        # redundancy check
        town_directory, month1, year1, *town = Scraper.parse_pdfname(pdfname)

        if pdfname in possible_corrupted_list:
            logger.info(f'PDF corrupted. The city of {town} for {month1} {year1} does not have data')
            Scraper.data_na(town, month1, year1)
            logger.removeHandler(f_handler)
            logger.removeHandler(c_handler)
            logging.shutdown()

        else:
            if update == 'Yes':
                os.chdir(f'C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files\\{year1}\\{" ".join(town_directory)}')
            else:
                os.chdir('C:\\Users\\Omar\\Desktop\\Python Temp Folder')
            try:
                with open(pdfname, 'rb') as reader:
                    pdfread = PyPDF2.PdfReader(reader)
                    page = pdfread.pages[0]
                    target = page.extract_text()
                    lines = target.split('\n')
                    lines = lines[24:]

                if type(town) is list:
                    try:
                        if town[0] == lines[2]:
                            real_town = town[0]
                            county = self.check_county(lines[3], real_town)

                        # If the town from the file name does not match the town name
                        # found inside the file @ first known location, check the second location
                        elif town[0] == lines[4]:
                            real_town = town[0]
                            county = self.check_county(lines[5], real_town)

                        # Check if any recognizable town name is found inside the target pdf location
                        else:
                            real_town, county = self.pdf_redundancy_check(lines, pdfname, month1, year1, logger)

                    except KeyError:
                        logger.exception(f'***{real_town} of {town[1]} is not in the state dictionary.\n'
                                         f'The program will use the county name supplied from the parsed_pdf function')
                        county = town[1]
                    finally:
                        assert county == town[1], f'{pdfname} corrupted. County names does not match'

                elif type(town) is not list:
                    if town == lines[2]:
                        real_town = town
                        county = self.check_county(lines[3], real_town)

                    elif town == lines[4]:
                        real_town = town
                        county = self.check_county(lines[5], real_town)

                    else:
                        real_town, county = self.pdf_redundancy_check(lines, pdfname, month1, year1, logger)

                Scraper.good_data(pdfname, target, real_town, county, month1, year1, main_dict, year_dict)
                logger.info(f'The data for {pdfname} has been extracted')

            except PyPDF2._reader.EmptyFileError as efe:
                logger.exception(f'An Error Has Occured (File Possibly Corrupted):\n{traceback.format_exception(efe)}')
                logger.info(f'The city of {real_town} for {month1} {year1} does not have data')
                # If function encounters an empty/corrupted pdf,
                # the data_na function will render all information available for that file equal to zero
                # The generator will then pick up at the next pdf to continue extracting data
                Scraper.data_na(real_town, month1, year1, main_dict, year_dict)

            except re.error as ree:
                logger.exception(f'A Regex Error Has Occurred:\n{traceback.format_exception(ree)}')

            except AssertionError as AE:
                logger.exception(f'An AssertionError Has Occurred:\n{traceback.format_exception(AE)}')
                logger.info(f'PDF corrupted. The city of {real_town} for {month1} {year1} does not have data')
                Scraper.data_na(real_town, month1, year1, main_dict, year_dict)

            except Exception as E:
                logger.exception(f'An Unhandled Error Has Occurred:\n{traceback.format_exception(E)}')

            finally:
                logger.removeHandler(f_handler)
                logger.removeHandler(c_handler)
                logging.shutdown()

    @logger_decorator
    def fill_missing_data(self, target_directories: list, main_dict, year_dict, **kwargs):
        """

        :param target_directories:
        :param main_dict:
        :param year_dict:
        :param kwargs:
        :return:
        """

        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']

        name = "Fill Missing Data"
        start_time = datetime.datetime.now()

        for pdf in Scraper.pdf_generator(pdfname=target_directories):
            self.extract_re_data(pdf, ['No Corrupted Files'], main_dict,year_dict, update='Yes')

        end_time = datetime.datetime.now()
        run_time = end_time - start_time

        self.event_log_update(name, run_time, logger)
        winsound.PlaySound('F:\\Python 2.0\\SoundFiles\\Victory.wav', 0)

        logger.removeHandler(f_handler)
        logger.removeHandler(c_handler)
        logging.shutdown()

    @staticmethod
    def find_closed_sales(pdf_text, month_var=None):

        variable_list = []

        closed_sales_pattern = re.compile(
            r'Closed\sSales\s(\d{0,3}?)\s(\d{0,3}?)\s(0.0%|--|[+-]\s\d{0,3}?.\d{0,1}?%)\s(\d{0,3}?)\s(\d{0,3}?)\s(0.0%|--|[+-]\s\d{0,3}?.\d{0,1}?%)')
        closed_sales_search = list(closed_sales_pattern.findall(pdf_text))
        closed_sales_current = int(closed_sales_search[0][1])
        closed_sales_pc = closed_sales_search[0][2].split(' ')
        closed_sales_per_change = ''.join(closed_sales_pc).rstrip('%')
        if '+' in closed_sales_per_change:
            closed_sales_per_change.lstrip('+')
            closed_sales_per_change = round(float(closed_sales_per_change) / 100, 3)
        elif '--' in closed_sales_per_change:
            closed_sales_per_change = 0.0
        else:
            closed_sales_per_change = round(float(closed_sales_per_change) / 100, 3)

        variable_list.extend([closed_sales_current, closed_sales_per_change])

        if month_var == 'December':

            closed_sales_fy = int(closed_sales_search[0][4])
            closed_sales_pc_fy = closed_sales_search[0][5].split(' ')
            closed_sales_per_change_fy = ''.join(closed_sales_pc_fy).rstrip('%')
            if '+' in closed_sales_per_change_fy:
                closed_sales_per_change_fy.lstrip('+')
                closed_sales_per_change_fy = float(closed_sales_per_change_fy) / 100
            elif '--' in closed_sales_per_change_fy:
                closed_sales_per_change_fy = 0.0
            else:
                closed_sales_per_change_fy = float(closed_sales_per_change_fy) / 100

            variable_list.extend([closed_sales_fy, closed_sales_per_change_fy])

        return variable_list


    @classmethod
    def find_county(cls, city):
        """
        Classmethod which returns the associated county of the arg "city"
        :param city: str variable of the name of a city or township
        :return: county name
        """

        return cls.state_dict.loc[city, 'County']

    @staticmethod
    def find_dom(pdf_text, month_var=None):

        variable_list = []

        dom_pattern = re.compile(
            r'Days\son\sMarket\sUntil\sSale\s(\d{0,3}?)\s(\d{0,3}?)\s(0.0%|--|[+-]\s\d{0,3}?.\d{0,1}?%)\s(\d{0,3}?)\s(\d{0,3}?)\s(0.0%|--|[+-]\s\d{0,3}?.\d{0,1}?%)')
        dom_search = list(dom_pattern.findall(pdf_text))
        dom_current = int(dom_search[0][1])
        dom_pc = dom_search[0][2].split(' ')
        dom_per_change = ''.join(dom_pc).rstrip('%')
        if '+' in dom_per_change:
            dom_per_change.lstrip('+')
            dom_per_change = round(float(dom_per_change) / 100, 3)
        elif '--' in dom_per_change:
            dom_per_change = 0.0
        else:
            dom_per_change = round(float(dom_per_change) / 100, 3)

        variable_list.extend([dom_current, dom_per_change])

        if month_var == 'December':
            dom_fy = int(dom_search[0][4])
            dom_pc_fy = dom_search[0][5].split(' ')
            dom_per_change_fy = ''.join(dom_pc_fy).rstrip('%')
            if '+' in dom_per_change_fy:
                dom_per_change_fy.lstrip('+')
                dom_per_change_fy = float(dom_per_change_fy) / 100
            elif '--' in dom_per_change_fy:
                dom_per_change_fy = 0.0
            else:
                dom_per_change_fy = float(dom_per_change_fy) / 100

            variable_list.extend([dom_fy, dom_per_change_fy])

        return variable_list

    @staticmethod
    def find_inventory(pdf_text, month_var=None):

        variable_list = []

        inventory_pattern = re.compile(
            r'Inventory\sof\sHomes\sfor\sSale\s(--|\d{0,3}?)\s(--|\d{0,3}?)\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)\s(--|\d{0,3}?)\s(--|\d{0,3}?)\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)')
        inventory_search = list(inventory_pattern.findall(pdf_text))
        inventory_current = inventory_search[0][1]
        if inventory_current != '--':
            inventory_current = int(inventory_current)
        inventory_pc = inventory_search[0][2].split(' ')
        inventory_per_change = ''.join(inventory_pc).rstrip('%')
        if '+' in inventory_per_change:
            inventory_per_change.lstrip('+')
            inventory_per_change = round(float(inventory_per_change) / 100, 3)
        elif '--' in inventory_per_change:
            inventory_per_change = 0.0
        else:
            inventory_per_change = round(float(inventory_per_change) / 100, 3)

        variable_list.extend([inventory_current, inventory_per_change])

        if month_var == 'December':

            inventory_fy = inventory_search[0][4]
            if inventory_fy == '--':
                inventory_fy = 0.0
            inventory_pc_fy = inventory_search[0][5].split(' ')
            inventory_per_change_fy = ''.join(inventory_pc_fy).rstrip('%')
            if '+' in inventory_per_change_fy:
                inventory_per_change_fy.lstrip('+')
                inventory_per_change_fy = float(inventory_per_change_fy) / 100
            elif '--' in inventory_per_change_fy:
                inventory_per_change_fy = 0.0
            else:
                inventory_per_change_fy = float(inventory_per_change_fy) / 100

            variable_list.extend([inventory_fy, inventory_per_change_fy])

        return variable_list

    @staticmethod
    def find_key_metrics(pdf_text):

        key_metrics_basic_pattern = re.compile(
            r'Key\sMetrics\s(\d{4})\s(\d{4})\sPercent\sChange\sThru\s\d{1,2}?-\d{4}\sThru\s\d{1,2}?-\d{4}\sPercent\sChange')
        km_search = list(key_metrics_basic_pattern.findall(pdf_text))

        return km_search[0][1]

    @staticmethod
    def find_median_sales(pdf_text, month_var=None):

        variable_list = []

        median_sales_pattern = re.compile(
            r'Median\sSales\sPrice\*\s(\$\d{1}|\$\d{0,3}?,?\d{0,3}?,\d{1,3})\s(\$\d{1}|\$\d{0,3}?,?\d{0,3}?,\d{1,3})\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)\s(\$\d{1}|\$\d{0,3}?,?\d{0,3}?,\d{1,3})\s(\$\d{1}|\$\d{0,3}?,?\d{0,3}?,\d{1,3})\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)')
        median_sales_search = list(median_sales_pattern.findall(pdf_text))
        median_sales_current = median_sales_search[0][1]
        median_sales_current = int("".join(median_sales_current.split(',')).lstrip('$'))
        median_sales_pc = median_sales_search[0][2].split(' ')
        median_sales_per_change = ''.join(median_sales_pc).rstrip('%')
        if '+' in median_sales_per_change:
            median_sales_per_change.lstrip('+')
            median_sales_per_change = round(float(median_sales_per_change) / 100, 3)
        elif '--' in median_sales_per_change:
            median_sales_per_change = 0.0
        else:
            median_sales_per_change = round(float(median_sales_per_change) / 100, 3)

        variable_list.extend([median_sales_current, median_sales_per_change])

        if month_var == 'December':

            median_sales_fy = median_sales_search[0][4]
            median_sales_fy = int("".join(median_sales_fy.split(',')).lstrip('$'))
            median_sales_pc_fy = median_sales_search[0][5].split(' ')
            median_sales_per_change_fy = ''.join(median_sales_pc_fy).rstrip('%')
            if '+' in median_sales_per_change_fy:
                median_sales_per_change_fy.lstrip('+')
                median_sales_per_change_fy = float(median_sales_per_change_fy) / 100
            elif '--' in median_sales_per_change_fy:
                median_sales_per_change_fy = 0.0
            else:
                median_sales_per_change_fy = float(median_sales_per_change_fy) / 100

            variable_list.extend([median_sales_fy, median_sales_per_change_fy])

        return variable_list

    @staticmethod
    def find_month(pdf_text):

        month_pattern = re.compile(
            r'(January|February|March|April|May|June|July|August|September|October|November|December)\sYear\sto\sDate\sSingle\sFamily')

        return month_pattern.search(pdf_text).group(1)

    @staticmethod
    def find_new_listings(pdf_text, month_var=None):

        variable_list = []

        new_listings_pattern = re.compile(
            r'New\sListings\s(\d{0,3}?)\s(\d{0,3}?)\s(0.0%|--|[+-]\s\d{0,3}?.\d{0,1}?%)\s(\d{0,3}?)\s(\d{0,3}?)\s(0.0%|--|[+-]\s\d{0,3}?.\d{0,1}?%)')
        new_listing_search = list(new_listings_pattern.findall(pdf_text))
        new_listings_current = int(new_listing_search[0][1])
        new_listings_pc = str(new_listing_search[0][2]).split(' ')
        new_listings_per_change = ''.join(new_listings_pc).rstrip('%')
        if '+' in new_listings_per_change:
            new_listings_per_change.lstrip('+')
            new_listings_per_change = round(float(new_listings_per_change) / 100, 3)
        elif '--' in new_listings_per_change:
            new_listings_per_change = 0.0
        else:
            new_listings_per_change = round(float(new_listings_per_change) / 100, 3)

        variable_list.extend([new_listings_current, new_listings_per_change])

        if month_var == 'December':

            new_listings_fy = int(new_listing_search[0][4])
            new_listings_pc_fy = str(new_listing_search[0][5]).split(' ')
            new_listings_per_change_fy = ''.join(new_listings_pc_fy).rstrip('%')
            if '+' in new_listings_per_change_fy:
                new_listings_per_change_fy.lstrip('+')
                new_listings_per_change_fy = float(new_listings_per_change_fy) / 100
            elif '--' in new_listings_per_change_fy:
                new_listings_per_change_fy = 0.0
            else:
                new_listings_per_change_fy = float(new_listings_per_change_fy) / 100

            variable_list.extend([new_listings_fy, new_listings_per_change_fy])

        return variable_list

    @staticmethod
    def find_percent_lpr(pdf_text, month_var=None):

        variable_list = []

        percent_lpr_pattern = re.compile(
            r'Percent\sof\sList\sPrice\sReceived\*\s(\d{1,3}?.\d{1}%)\s(\d{1,3}?.\d{1}%)\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)\s(\d{1,3}?.\d{1}%)\s(\d{1,3}?.\d{1}%)\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)')
        percent_lpr_search = list(percent_lpr_pattern.findall(pdf_text))
        # Divide this by 100 and figure out how to format these to show the percent sign
        percent_lpr_current = float(percent_lpr_search[0][1].rstrip('%'))
        percent_lpr_pc = percent_lpr_search[0][2].split(' ')
        percent_lpr_per_change = ''.join(percent_lpr_pc).rstrip('%')
        if '+' in percent_lpr_per_change:
            percent_lpr_per_change.lstrip('+')
            percent_lpr_per_change = round(float(percent_lpr_per_change) / 100, 3)
        elif '--' in percent_lpr_per_change:
            percent_lpr_per_change = 0.0
        else:
            percent_lpr_per_change = round(float(percent_lpr_per_change) / 100, 3)

        variable_list.extend([percent_lpr_current, percent_lpr_per_change])

        if month_var == 'December':

            percent_lpr_fy = float(percent_lpr_search[0][4].rstrip('%'))
            percent_lpr_pc_fy = percent_lpr_search[0][5].split(' ')
            percent_lpr_per_change_fy = ''.join(percent_lpr_pc_fy).rstrip('%')
            if '+' in percent_lpr_per_change_fy:
                percent_lpr_per_change_fy.lstrip('+')
                percent_lpr_per_change_fy = float(percent_lpr_per_change_fy) / 100
            elif '--' in percent_lpr_per_change_fy:
                percent_lpr_per_change_fy = 0.0
            else:
                percent_lpr_per_change_fy = float(percent_lpr_per_change_fy) / 100

            variable_list.extend([percent_lpr_fy, percent_lpr_per_change_fy])

        return variable_list

    @staticmethod
    def find_supply(pdf_text, month_var=None):

        variable_list = []

        supply_pattern = re.compile(
            r'Months\sSupply\sof\sInventory\s(--|\d{1,2}?.\d{1})\s(--|\d{1,2}?.\d{1})\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)\s(--|\d{1,2}?.\d{1})\s(--|\d{1,2}?.\d{1})\s(0.0%|--|[+-]\s\d{1,3}?.\d{1}%)')
        supply_search = list(supply_pattern.findall(pdf_text))
        supply_current = supply_search[0][1]
        if supply_current != '--':
            supply_current = float(supply_current)
        supply_pc = supply_search[0][2].split(' ')
        supply_per_change = ''.join(supply_pc).rstrip('%')
        if '+' in supply_per_change:
            supply_per_change.lstrip('+')
            supply_per_change = round(float(supply_per_change) / 100, 3)
        elif '--' in supply_per_change:
            supply_per_change = 0.0
        else:
            supply_per_change = round(float(supply_per_change) / 100, 3)

        variable_list.extend([supply_current, supply_per_change])

        if month_var == 'December':

            supply_fy = supply_search[0][4]
            if supply_fy == '--':
                supply_fy = 0.0
            supply_pc_fy = supply_search[0][5].split(' ')
            supply_per_change_fy = ''.join(supply_pc_fy).rstrip('%')
            if '+' in supply_per_change_fy:
                supply_per_change_fy.lstrip('+')
                supply_per_change_fy = float(supply_per_change_fy) / 100
            elif '--' in supply_per_change_fy:
                supply_per_change_fy = 0.0
            else:
                supply_per_change_fy = float(supply_per_change_fy) / 100

            variable_list.extend([supply_fy, supply_per_change_fy])

        return variable_list

    # Function which pulls the username and password for a specified website
    @staticmethod
    def get_us_pw(website):
        """
        Staticmethod which fetches the username and passord for the arg "website"
        :param website: Name of the website for which the username and password are needed
        :return: username, pw
        """
        # Saves the current directory in a variable in order to switch back to it once the program ends
        previous_wd = os.getcwd()
        os.chdir('F:\\Jibreel Hameed\\Kryptonite')
        wb = openpyxl.load_workbook('get_us_pw.xlsx')
        sheet = wb.active
        for i,n in enumerate(sheet['A0' : 'A20']):
            for cell in n:
                if website == cell.value:
                    username = sheet['C' + str(i+1)].value
                    pw = sheet['D' + str(i+1)].value

        os.chdir(previous_wd)

        return username, pw

    @staticmethod
    def good_data(pdfname, target, city, county, month1, year1, main_dict, year_dict):
        """

        :param pdfname:
        :param target:
        :param city:
        :param county:
        :param month1:
        :param year1:
        :param main_dict:
        :param year_dict:
        :return:
        """

        month = Scraper.find_month(target)
        quarter = Scraper.quarter(month)
        current_year = Scraper.find_key_metrics(target)
        closed_sales_list = Scraper.find_closed_sales(target, month)
        closed_sales_current = closed_sales_list[0]
        closed_sales_per_change = closed_sales_list[1]
        dom_list = Scraper.find_dom(target, month)
        dom_current = dom_list[0]
        dom_per_change = dom_list[1]
        inventory_list = Scraper.find_inventory(target, month)
        inventory_current = inventory_list[0]
        inventory_per_change = inventory_list[1]
        median_sales_list = Scraper.find_median_sales(target, month)
        median_sales_current = median_sales_list[0]
        median_sales_per_change = median_sales_list[1]
        new_listings_list = Scraper.find_new_listings(target, month)
        new_listings_current = new_listings_list[0]
        new_listings_per_change = new_listings_list[1]
        per_lpr_list = Scraper.find_percent_lpr(target, month)
        percent_lpr_current = per_lpr_list[0]
        percent_lpr_per_change = per_lpr_list[1]
        supply_list = Scraper.find_supply(target, month)
        supply_current = supply_list[0]
        supply_per_change = supply_list[1]

        category_list = ['City', 'County', 'Quarter', 'Month', 'Year', 'New Listings',
                         'New Listing % Change (YoY)', 'Closed Sales',
                         'Closed Sale % Change (YoY)', 'Days on Markets', 'Days on Market % Change (YoY)',
                         'Median Sales Prices',
                         'Median Sales Price % Change (YoY)', 'Percent of Listing Price Received',
                         'Percent of Listing Price Receive % Change (YoY)', 'Inventory of Homes for Sales',
                         'Inventory of Homes for Sale % Change (YoY)', 'Months of Supply',
                         'Months of Supplies % Change (YoY)']

        variable_list = [city, county, quarter, month, current_year, new_listings_current,
                         new_listings_per_change, closed_sales_current,
                         closed_sales_per_change, dom_current, dom_per_change, median_sales_current,
                         median_sales_per_change, percent_lpr_current, percent_lpr_per_change,
                         inventory_current, inventory_per_change, supply_current, supply_per_change]

        # First check to see if the pdf contains the correct data
        assert month == month1 and current_year == year1, f'{pdfname} is corrupted. ' \
                                                          f'Giving data for {city} {month} {current_year}.pdf'

        if main_dict[current_year] == {}:
            for idx, n in enumerate(category_list):
                main_dict[current_year].setdefault(n, [])
                main_dict[current_year][n].append(variable_list[idx])
        else:
            # Redundancy check to make sure there isn't any duplicate vectors in the database
            Scraper.duplicate_vector_check(pdfname, variable_list, current_year, main_dict)
            for idx, n in enumerate(category_list):
                main_dict[current_year][n].append(variable_list[idx])

        if month == 'December':
            category_list1 = category_list[:]
            del category_list1[2]

            new_listings_fy = new_listings_list[2]
            new_listings_per_change_fy = new_listings_list[3]
            closed_sales_fy = closed_sales_list[2]
            closed_sales_per_change_fy = closed_sales_list[3]
            dom_fy = dom_list[2]
            dom_per_change_fy = dom_list[3]
            median_sales_fy = median_sales_list[2]
            median_sales_per_change_fy = median_sales_list[3]
            percent_lpr_fy = per_lpr_list[2]
            percent_lpr_per_change_fy = per_lpr_list[3]
            inventory_fy = inventory_list[2]
            inventory_per_change_fy = inventory_list[3]
            supply_fy = supply_list[2]
            supply_per_change_fy = supply_list[3]

            fy_variable_list = [city, county, month, current_year, new_listings_fy,
                                new_listings_per_change_fy, closed_sales_fy, closed_sales_per_change_fy,
                                dom_fy, dom_per_change_fy, median_sales_fy, median_sales_per_change_fy,
                                percent_lpr_fy, percent_lpr_per_change_fy, inventory_fy,
                                inventory_per_change_fy, supply_fy, supply_per_change_fy]

            if year_dict[current_year] == {}:
                for idx, n in enumerate(category_list1):
                    year_dict[current_year].setdefault(n, [])
                    year_dict[current_year][n].append(fy_variable_list[idx])
            else:
                for idx, n in enumerate(category_list1):
                    year_dict[current_year][n].append(fy_variable_list[idx])

    # Used in case the njr10k or the update_njr10k functions are used recursively.
    # This function will find the latest file downloaded and continue from that point
    def latest_file(self):

        base_path = 'C:\\Users\\Omar\\Desktop\\Python Temp Folder'

        for root, dirs, filenames in os.walk(base_path):
            # print(filenames)
            # If the length of the filenames variable is less than 2,
            # that means there are no PDF Files in the folder. The 2 existing files are log files
            if len(filenames) <= 2:
                return 'None'
            else:
                try:
                    i = -1
                    while not filenames[i].endswith('.pdf'):
                        # Search the filenames list in reverse order to look for the latest PDF file
                        i -= 1
                    else:
                        target = filenames[i]
                        # Files are stored in alphabetical order
                        # Wyckoff Twp are the last pdfs to be downloaded
                        # If the target pdf name is either equal to Wyckoff Twp Sept for this year or last year
                        # or Wyckoff Twp of the current month and year
                        # All files may be downloaded, need to do a second check
                        check1 = 'Wyckoff Twp September ' + str(int(Scraper.current_data.split(' ')[1]) - 1) + '.pdf'
                        check2 = 'Wyckoff Twp September ' + Scraper.current_data.split(' ')[1] + '.pdf'
                        check3 = 'Wyckoff Twp ' + Scraper.current_data + '.pdf'
                        if target == check1 or target == check2 or target == check3:
                            print(f'Latest file downloaded is: {target}\nAll files may be downloaded...')
                            time.sleep(1)
                            print(f'Moving 2nd phase check...')
                            time.sleep(1)
                        else:
                            print(f'Latest file downloaded is: {target}\nRestarting download process...')
                            info = target.rstrip('.pdf').split(' ')
                            town = info[0:len(info) - 2]
                            if len(town) > 1:
                                town = ' '.join(town)
                            else:
                                town = info[0]

                except IndexError:
                    # The code tried searching outside the list because it didn't find a PDF file
                    print(f'There are currently no downloaded PDFs available. Starting NJR10k download...')
                    return 'None'
                else:
                    check4 = 'Wyckoff Twp ' + Scraper.current_data + '.pdf'
                    if os.path.exists(base_path + '\\' + check4):
                        print('All files have been downloaded. Now reading previous logger file...')
                        return 'Read Logger File'
                    else:
                        full_town_list = self.__towns
                        return full_town_list[full_town_list.index(town):]

    @staticmethod
    @logger_decorator
    def matplot_lines(filename, **kwargs):

        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']

        by_qtr_stats = pd.read_excel(filename, sheet_name='All Months')
        # Need to use tolist() vs to_list() because the latter creates a numpy array first
        years = by_qtr_stats['Year'].unique().tolist()
        temp_df = by_qtr_stats.copy().sort_values(by='Dates')
        counties = temp_df['County'].unique().tolist()
        target_columns = ['New Listings', 'Closed Sales', 'Days on Markets', 'Median Sales Prices',
                          'Percent of Listing Price Received']
        # Dynamically numbering the Figures to not overlap the graphs and labels for each county
        fig1_num = 1

        for idx, county in enumerate(counties):

            temp_df1 = temp_df[temp_df['County'] == county]
            cities = temp_df1['City'].unique().tolist()
            cols = math.ceil(len(cities) / 10)

            pdfname = 'F:\\Real Estate Investing\\JQH Holding Company LLC\\Property Data\\' + county + ' Stats PDF.pdf'
            with PdfPages(pdfname) as PDFMaker:
                for city in cities:
                    plot_container = []
                    lab_container = []
                    # The number of columns, rows and figure numbers have to positive integers greater than 0
                    plt.figure(fig1_num, layout='constrained', figsize=(11, 6))
                    plt.suptitle('Statistical Data By Quarter for ' + county)
                    for idx1, column in enumerate(target_columns):
                        target_df = temp_df1[temp_df1['City'] == city]
                        plt.figure(fig1_num)
                        axs = plt.subplot(3, 2, idx1 + 1)
                        Scraper.nj10k_linechart_plotter(axs)
                        plt.plot(target_df['Dates'], target_df[column], label=city)
                        plt.plot(target_df['Dates'], [target_df[column].mean() for _ in range(len(target_df['Dates']))],
                                 color='black', label=f'{len(years)} Year Avg')
                        plt.ylabel(column, fontsize='small')

                    plot, label = axs.get_legend_handles_labels()
                    plot_container.extend(plot)
                    lab_container.extend(label)
                    fig1 = plt.figure(fig1_num)
                    plt.figlegend(plot_container, lab_container, loc='outside left upper', bbox_to_anchor=(0.53, 0.32),
                                  ncols=cols, title='Cities in ' + county, fontsize='x-small', title_fontsize='x-small',
                                  markerscale=0.4)
                    PDFMaker.savefig(fig1)
                    plt.close()

                    fig1_num += 1

    # Function which logs into njrealtor to automatically download the pdfs from each city to get ready to scrape
    @logger_decorator
    def njr10k(self, **kwargs):

        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']
        start_time = datetime.datetime.now()

        name = 'NJR10k'

        possible_corrupted_files = []

        base_url = 'http://njar.stats.10kresearch.com/docs/lmu/'

        # Initiates a Requests Session which maintains the cookies and session info until closed
        # This is needed in order to successfully log into njrealtor and access the 10k pdfs
        with requests.Session() as session:
            # create a function/module which returns the njr10k info
            username, pw = Scraper.get_us_pw('NJRealtor')

            # payload sent during the HTTP POST
            payload1 = {'rd': '10',
                        'passedURL': '/goto/10k/',
                        'case': '',
                        'LoginEmail': username,
                        'LoginPassword': pw,
                        'LoginButton': 'Login'}

            # website parameters needs to access the correct pdf addresses
            params = {'src': 'Page'}
            months = list(self.__months.keys())

            url = 'https://www.njrealtor.com/login/?rd=10&passedURL=/goto/10k/'
            url2 = 'https://www.njrealtor.com/ramco-api/web-services/login_POST.php'

            session.get(url)  # Request to arrive at the log-in page
            session.post(url2, data=payload1)  # Response object to logging into website

            # If this is a recursive run, towns_list will be a sliced list starting from the last run city
            # Latest_file searches for the last
            towns_list = self.latest_file()

            if towns_list == 'None':
                towns_list = self.__towns
            elif towns_list == 'Read Logger File':
                return 'Read Logger File'

            try:
                for i in towns_list:
                    # Takes the name of the city from the list and splits the string at the space,
                    # then joins the strings in the newly created list
                    # This is needed to use in the url3 variable to access the correct 10k pdfs

                    for y in self.__years:

                        if y == '2019':
                            # If year = 2019, there is no data available from January to Sept.
                            # Also some cities will not have data available for Sept and produce data for other dates
                            months1 = months[8:13]
                            for m in months1:

                                url3, new_filename = self.create_url_and_pdfname(base_url, y, m, i)

                                Scraper.download_pdf(session, url3, params, new_filename, possible_corrupted_files,
                                                     logger)

                        elif y == self.__years[-1]:
                            # If year is the latest year, months1 will equal a sliced list of the
                            # first month represented by 01 to the latest month represent by 2 digits
                            for k, v in self.__months.items():
                                # If v equals the month of the most current data
                                if v == Scraper.current_data.split(' ')[0]:
                                    months1 = months[:months.index(k) + 1]
                            for m in months1:
                                url3, new_filename = self.create_url_and_pdfname(base_url, y, m, i)

                                Scraper.download_pdf(session, url3, params, new_filename, possible_corrupted_files,
                                                     logger)

                        elif y != '2019':
                            for m in months:
                                url3, new_filename = self.create_url_and_pdfname(base_url, y, m, i)

                                Scraper.download_pdf(session, url3, params, new_filename, possible_corrupted_files,
                                                     logger)

            except IOError:
                """An OS Error has occurred """
                logger.exception(f'IOError has Occurred')
                self.njr10k()

            except requests.exceptions.HTTPError as rht:
                """An HTTP error occurred."""
                logger.exception(f'An HTTP has Occurred: {rht}')

            except requests.exceptions.Timeout as ret:
                """The request timed out.
    
                Catching this error will catch both
                :exc:`~requests.exceptions.ConnectTimeout` and
                :exc:`~requests.exceptions.ReadTimeout` errors.
                """
                logger.exception(f'The Request Has Timed Out: {ret}')

            except requests.exceptions.InvalidURL as rei:
                """The URL provided was somehow invalid."""
                logger.exception(f'The URL Provided Was Invalid: {rei}')

            except requests.exceptions.RetryError as rer:
                """Custom retries logic failed"""
                logger.exception(f'Custom Retries Logic Failed: {rer}')

            except requests.exceptions.StreamConsumedError as res:
                """The content for this response was already consumed."""
                logger.exception(f'The Content For This Response Was Already Consumed: {res}')

            except requests.exceptions.ContentDecodingError as rec:
                """Failed to decode response content."""
                logger.exception(f'Failed to Decode Response Content: {rec}')

            except requests.exceptions.ChunkedEncodingError as rece:
                """The server declared chunked encoding but sent an invalid chunk."""
                logger.exception(f'Invalid Chunk Encoding: {rece}')

            except Exception as e:
                logger.exception(f'An Unhandled Error Has Occurred: {e}')

            else:

                end_time = datetime.datetime.now()
                run_time = end_time - start_time

                self.event_log_update(name, run_time, logger)
                winsound.PlaySound('F:\\Python 2.0\\SoundFiles\\Victory.wav', 0)

                logger.removeHandler(f_handler)
                logger.removeHandler(c_handler)
                logging.shutdown()

        if len(possible_corrupted_files) > 0:
            return possible_corrupted_files
        else:
            return "All Files Downloaded"

    @staticmethod
    def nj10k_linechart_plotter(axes_label):
        axes_label.tick_params(axis='y', labelsize='x-small')
        axes_label.tick_params(axis='x', labelsize='x-small', labelrotation=45)
        axes_label.xaxis.set_major_locator(mdates.MonthLocator(bymonth=(1, 4, 7, 10)))
        axes_label.xaxis.set_minor_locator(mdates.MonthLocator())
        axes_label.yaxis.set_minor_locator(ticker.AutoMinorLocator())
        axes_label.xaxis.set_major_formatter(mdates.DateFormatter('%y-%b'))

    # Function uses Selenium to webscrape the cities and counties from the njrealtor 10k website
    @logger_decorator
    def njrdata(self, **kwargs):

        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']
        options = Options()
        # Change this directory to the new one: ('C:\\Users\\Omar\\Desktop\\Python Temp Folder')
        s = {"savefile.default_directory": 'C:\\Users\\Omar\\Desktop\\Selenium Temp Folder'}
        # options.add_argument('window-postion=2000,0')
        # options.add_experimental_option("detach", True)
        options.add_experimental_option("prefs", s)
        options.add_argument("--headless=new")
        driver = webdriver.Edge(service=Service(), options=options)
        # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        # driver = webdriver.Chrome(service=Service
        # (ChromeDriverManager(version='114.0.5735.90').install()), options = options)
        url = 'https://www.njrealtor.com/login.php?rd=10&passedURL=/goto.php?10kresearch=1&skipToken=1'
        driver.get(url)

        username, pw = Scraper.get_us_pw('NJRealtor')

        try:
            # Login in using my email and password
            email = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='LoginEmail']"))
            )
            email.send_keys(username)
            pw1 = driver.find_element(By.XPATH, "//input[@id='LoginPassword']")
            pw1.send_keys(pw)
            login = driver.find_element(By.XPATH, "//input[@id='LoginButton']")
            login.click()

            # Recognize the page element to know it's time to webscrape all the cities and counties
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//img[@class='brand']"))
            )
            page_results = driver.page_source
            soup = BeautifulSoup(page_results, 'html.parser')
            self.area_results(soup)
            self.current_data_avail(soup)

        except TimeoutException as te:
            logger.exception(f'Timeout Error Occurred: {te}')

        except NoSuchElementException as nse:
            logger.exception(f'So Such Element Was Found: {nse}')

        except ElementNotVisibleException as env:
            logger.exception(f'The Element Is Not Visible: {env}')

        except Exception as e:
            logger.exception(f'An Unhandled Error Has Occurred: {e} ')

        else:
            logger.info(f'BeautifulSoup Has Run for NJ Realtor Successfully')
            logger.removeHandler(f_handler)
            logger.removeHandler(c_handler)
            logging.shutdown()

    # Function which will organize the PDF Temp Files Folder by year, month and city
    @staticmethod
    def OrganizeFiles(results_from_corrupt):
        print('Now organizing files into folders separated by year and city...')
        base_path = 'C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files'
        target_path = 'C:\\Users\\Omar\\Desktop\\Python Temp Folder'

        os.chdir(target_path)
        filenames = os.listdir(target_path)
        # for filenames in os.path(target_path):
        # If the first index of the list is 'No Corrupted Files', the send2trash function isn't necessary
        if results_from_corrupt[0] == 'No Corrupted Files':
            for filename in filenames:
                if filename.endswith('.pdf'):
                    target = filename.rstrip('.pdf').split(' ')
                    year = target[-1]
                    # month = target[-2]
                    city = ' '.join(target[:len(target) - 2])

                    first_path = os.path.join(base_path, year)
                    second_path = os.path.join(first_path, city)
                    abs_file_path = os.path.abspath(filename)
                    # If the first target path exists, check to see if the city path exists as well
                    if os.path.exists(first_path):

                        if os.path.exists(second_path):
                            shutil.move(abs_file_path, second_path)

                        elif not os.path.exists(second_path):
                            os.makedirs(second_path)
                            shutil.move(abs_file_path, second_path)

                    elif not os.path.exists(first_path):
                        os.makedirs(second_path)
                        shutil.move(abs_file_path, second_path)
                else:
                    continue
        else:
            for filename in filenames:
                # If the filename is in the List, it is labeled corrupted and sent to the Trash Bin
                if filename in results_from_corrupt:
                    send2trash(filename)
                else:
                    if filename.endswith('.pdf'):
                        target = filename.rstrip('.pdf').split(' ')
                        year = target[-1]
                        month = target[-2]
                        city = ' '.join(target[:len(target) - 2])

                        first_path = os.path.join(base_path, year)
                        second_path = os.path.join(first_path, city)
                        abs_file_path = os.path.abspath(filename)
                        # If the first target path exists, check to see if the city path exists as well
                        if os.path.exists(first_path):

                            if os.path.exists(second_path):
                                shutil.move(abs_file_path, second_path)

                            elif not os.path.exists(second_path):
                                os.makedirs(second_path)
                                shutil.move(abs_file_path, second_path)

                        elif not os.path.exists(first_path):
                            os.makedirs(second_path)
                            shutil.move(abs_file_path, second_path)
                    else:
                        continue
        print('All files have been organized...')

    @logger_decorator
    def pandas2excel(self, dict1, dict2, **kwargs):

        """ The dict argument will be a nested dictionary.
              dict = {'2019' : {'New Listings': [],
                                'Closed Sales' : [],
                                etc}"""
        print('Storing Quarterly and Full Year Data into an Excel Spreadsheet...')
        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']
        previous_dir = os.getcwd()
        os.chdir('F:\\Real Estate Investing\\JQH Holding Company LLC\\Real Estate Data')
        # Store the dfs in one Excel file under different sheets for later processing
        filename = 'NJ 10k Real Estate Data ' + str(datetime.datetime.today().date()) + '.xlsx'
        with pd.ExcelWriter(filename) as writer:
            logger.info(f'Now creating Dataframes for Main Dictionary and Full Year')
            list1 = []
            list2 = []
            for k in dict1.keys():
                if k == '2018':
                    continue
                else:
                    df = pd.DataFrame(dict1[k])
                    convert_dict = {'Year': str}
                    df = df.astype(convert_dict)
                    df = df.assign(Dates=df['Month'] + df['Year'])
                    df['Dates'] = pd.to_datetime(df['Dates'].tolist(), format="%B%Y", errors='ignore')
                    df.insert(1, 'Dates', df.pop('Dates'))
                    df['Percent of Listing Price Received'] = df['Percent of Listing Price Received'] / 100
                    logger.info(f'Quarterly dataframe for {k} has been created')
                    list1.append(df)
                    df1 = pd.DataFrame(dict2[k])
                    logger.info(f'Yearly dataframe for {k} has been created')
                    list2.append(df1)
                    # df.to_excel(writer, sheet_name= k + ' By Month', index_label='City', merge_cells=False)
                    # df1.to_excel(writer, sheet_name= k + ' Full Year', index_label='City', merge_cells=False)
                    # logger.info(f'Both dataframes for {k} have been sent to the Excel file')

            logger.info(f'Now joining all quarterly dataframes...')
            all_months = pd.concat(list1)
            all_months.to_excel(writer, sheet_name='All Months', index_label='City', merge_cells=False)
            logger.info(f'All quarterly dataframes have been joined and sent to the Excel file')
            logger.info(f'Now joining all yearly dataframes...')
            all_years = pd.concat(list2)
            all_years.to_excel(writer, sheet_name='All Years', index_label='City', merge_cells=False)
            logger.info(f'All yearly dataframes have been joined and sent to the Excel file')

        os.chdir(previous_dir)
        print('Quarterly and Full Year dataframes have been created and stored in Excel\n'
              f'Filename: NJ 10k Real Estate Data {str(datetime.datetime.today().date())}')

        winsound.PlaySound('F:\\Python 2.0\\SoundFiles\\Victory.wav', 0)

        return filename

    @staticmethod
    def parse_pdfname(pdf_name):

        info = pdf_name.rstrip('.pdf').split(' ')
        town_directory = info[0:len(info) - 2]

        if len(town_directory) > 2:
            if 'County' in town_directory:
                # This means the city name is a duplicate and needs to have the county distinguished
                county = ' '.join(town_directory[-2:])
                town = ' '.join(town_directory[0:(town_directory.index('County') - 1)])
            else:
                town = ' '.join(town_directory)
        else:
            town = ' '.join(town_directory)

        month1 = info[-2]
        year1 = info[-1]

        if county:
            parsed_results = (town_directory, month1, year1, town, county)
        else:
            parsed_results = (town_directory, month1, year1, town)

        return parsed_results

    # Generator function which will be used in tandem with the extract_re_data function to put data into main dictionary
    @staticmethod
    def pdf_generator(pdfname=None):

        base_path = 'C:\\Users\\Omar\\Desktop\\Python Temp Folder'
        # Variable pdfname will either be a string argument or None
        pdfname = pdfname
        filenames = os.listdir(base_path)
        base_dir = os.getcwd()
        # for filenames in os.listdir(base_path):
        if pdfname is None:
            # If pdfname is None, the generator starts at the beginning of the list
            for filename in filenames:
                if filename.endswith('.pdf'):
                    yield filename
                else:
                    continue

        elif type(pdfname) is list:
            filenames = []
            years = ['2019', '2020', '2021', '2022', '2023']
            for year in years:
                for municipality in pdfname:
                    search_directory = f'C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files\\{year}\\{municipality}'
                    # os.chdir(search_directory)
                    missing_files = os.listdir(search_directory)
                    filenames.extend(missing_files)

            os.chdir(base_path)
            for filename in filenames:
                if filename.endswith('.pdf'):
                    yield filename
                else:
                    continue

        elif pdfname.endswith('.pdf'):
            # If pdfname is a string and ends with '.pdf', an error occurred in the extract_re_data function
            # The program will truncate the filesnames list starting with the first file name after the error
            filenames = filenames[filenames.index(pdfname) + 1:]
            for filename in filenames:
                if filename.endswith('.pdf'):
                    yield filename
                else:
                    continue

    def pdf_redundancy_check(self, pdf_text, pdf_name, month_var, year_var, logger):

        temp_town = None
        temp_county = None

        for i in self.__towns:
            if i not in pdf_text:
                continue
            else:
                temp_town = i

        for k in self.__counties:
            if k not in pdf_text:
                continue
            else:
                temp_county = k

        if temp_town or temp_county is None:
            logger.info(f'{pdf_name} corrupted. Does not have reliable data')
            Scraper.data_na(temp_town, month_var, year_var)
            raise Exception(f'{pdf_name} corrupted')

        elif os.path.exists(f'C:\\Users\\Omar\\Desktop\\Python Temp Folder\\' +
                            f'PDF Temp Files\\{year_var}\\{temp_town}'):
            # This logic will need to be changed in the future because the Python Temp Folder directory
            # gets zipped and moved at the end of the year. Need a permanent directory to check
            town = temp_town
            county = temp_county

            return town, county

        else:
            logger.info(f'{pdf_name} corrupted. The city of {temp_town} for {month_var} {year_var} does not have data')
            Scraper.data_na(temp_town, month_var, year_var)
            raise Exception(f'{pdf_name} corrupted')

    @staticmethod
    def quarter(month):
        if month in ['January', 'February', 'March']:
            quarter = 'Q1'
        elif month in ['April', 'May', 'June']:
            quarter = 'Q2'
        elif month in ['July', 'August', 'September']:
            quarter = 'Q3'
        elif month in ['October', 'November', 'December']:
            quarter = 'Q4'

        return quarter

    @staticmethod
    @run_quarterly_statistics(datetime.datetime.today())
    def quarterly_statistics(filename):
        # This method will run matplot_lines and njr10k_stats/njr10k_update_stats, cloropleth, histograms and Seaborn
        pass

    # A function I can use for debugging purposes.
    # In the event a failure occurs before reaching the extract_re_data function
    # I can read the logger file and start again as a midway point
    def read_logger(self, logger):

        base_path = 'C:\\Users\\Omar\\Desktop\\Python Temp Folder'
        corrupt_pattern = re.compile(r'\d+-\w+-\d+\s\d+:\d+:\d+\s-\supdate_njr10k|njr10k\s-\sWARNING\s-\sWARNING!\s(.*.pdf)\sis\spossibly\sa\scorrupted\sfile')
        run_name_pattern = re.compile(r'Run Type:\s(\w+\s\w+)')
        run_time_pattern = re.compile(r"Run Time:\s(.*)")

        possible_corrupted_files = []

        for root, dirs, filenames in os.walk(base_path):

            if len(filenames) < 1:
                return 'None'
            elif len(filenames) > 1:
                i = -1
                # Will find all log files
                while not filenames[i].endswith('.log'):
                    i -= 1
                else:
                    file_pattern = re.compile(r'update_njr10k|njr10k\s\d{4}-\d{2}-\d{2}.log')
                    while not file_pattern.search(filenames[i]):
                        i -= 1
                    else:
                        target = []
                        while len(target) < 1:
                            file = filenames[i]
                            with open('C:\\Users\\Omar\\Desktop\\Python Temp Folder\\' + file, 'r') as reader:
                                target = reader.readlines()
                                i -= 1
                        else:
                            break

        for j in target:
            if corrupt_pattern.search(j):
                if corrupt_pattern.search(j).group(1) is not None:
                    possible_corrupted_files.append(corrupt_pattern.search(j).group(1))
            elif run_name_pattern.search(j):
                name = run_name_pattern.search(j).group(1)
            elif run_time_pattern.search(j):
                run_time = run_time_pattern.search(j).group(1)
            else:
                continue

        self.event_log_update(name, run_time, logger)

        return possible_corrupted_files

    @classmethod
    def state_dictionary(cls):
        filename = 'F:\\Real Estate Investing\\JQH Holding Company LLC\\Real Estate Data\\NJ 10k Real Estate Data 2023-08-26.xlsx'
        df = pd.read_excel(filename, sheet_name='2022 Full Year', index_col=1)
        state_dict = df[['County']]

        cls.state_dict = state_dict

    @staticmethod
    def text_message(message_body):
        # Your Account SID from twilio.com/console
        account_sid = 'AC91ccb829e7e47ff05e69d8f96d627f73'
        # Your Auth Token from twilio.com/console
        auth_token = "f686fc30c3c9bc2694f11bb5137bb28c"

        twilioclient = Client(account_sid, auth_token)

        message = twilioclient.messages.create(
            to="+19084683728",
            from_="+19088609446",
            body=message_body)

        print(message.body)

    # This function can only be run after njr10k runs for the first time
    # This function will compare the last key-value pair of the class event_log variable
    # and the results of the current_data_avail function. If the values are not the same, run the program
    @logger_decorator
    def update_njr10k(self, start, finish, **kwargs):

        start_time = datetime.datetime.now()

        name = 'Update NJR10k'

        possible_corrupted_files = []
        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']

        base_url = 'http://njar.stats.10kresearch.com/docs/lmu/'

        # Initiates a Requests Session which maintains the cookies and session info until closed
        # This is needed in order to successfully log into njrealtor and access the 10k pdfs
        with requests.Session() as session:
            # create a function/module which returns the njr10k info
            username, pw = Scraper.get_us_pw('NJRealtor')

            # payload sent during the HTTP POST
            payload1 = {'rd': '10',
                        'passedURL': '/goto/10k/',
                        'case': '',
                        'LoginEmail': username,
                        'LoginPassword': pw,
                        'LoginButton': 'Login'}

            # website parameters needs to access the correct pdf addresses
            params = {'src': 'Page'}
            months = list(self.__months.keys())

            url = 'https://www.njrealtor.com/login/?rd=10&passedURL=/goto/10k/'
            url2 = 'https://www.njrealtor.com/ramco-api/web-services/login_POST.php'

            session.get(url)
            session.post(url2, data=payload1)

            start_month = start[0]
            start_year = start[1]
            current_month = finish[0]
            current_year = finish[1]

            # Assures that I'll have the correct year range when I slice the self.__years list
            assert int(start_year) <= int(current_year), "Invalid Operation: Start Year is greater than Current Year"

            # If the start_year and current year variables are the same,
            # form a one-object list to iterate through consisting of the current_year
            # Else, create a new year list which is the full range
            # from the start year to current year by slicing self.__year
            if start_year == current_year:
                years = [current_year]
            else:
                years = self.__years[self.__years.index(start_year): self.__years.index(current_year) + 1]

            for k, v in self.__months.items():
                if start_month == v:
                    start_month1 = k
            for k, v in self.__months.items():
                if current_month == v:
                    current_month1 = k
                    if os.path.exists('C:\\Users\\Omar\\Desktop\\Python Temp Folder\\PDF Temp Files\\'
                                      + current_year + '\\Wyckoff Twp\\Wyckoff Twp ' + start_month + ' '
                                      + current_year + '.pdf'):
                        start_month1 = current_month1

            # If this is a recursive run, towns_list will be a sliced list starting from the last run city
            towns_list = self.latest_file()

            if towns_list == 'None':
                towns_list = self.__towns
            elif towns_list == 'Read Logger File':
                return 'Read Logger File'

            for i in towns_list:

                # Takes the name of the city from the list and splits the string at the space,
                # then joins the strings in the newly created list
                # This is needed to use in the url3 variable to access the correct 10k pdfs

                for y in years:

                    months1 = months[months.index(start_month1): months.index(current_month1) + 1]
                    for m in months1:

                        url3, new_filename = self.create_url_and_pdfname(base_url, y, m, i)

                        Scraper.download_pdf(session, url3, params, new_filename, possible_corrupted_files,
                                             logger)

            end_time = datetime.datetime.now()
            run_time = end_time - start_time

            self.event_log_update(name, run_time, logger)
            winsound.PlaySound('F:\\Python 2.0\\SoundFiles\\Victory.wav', 0)

            logger.removeHandler(f_handler)
            logger.removeHandler(c_handler)
            logging.shutdown()

        if len(possible_corrupted_files) > 0:
            return possible_corrupted_files
        else:
            return "All Files Downloaded"

    @staticmethod
    def waiting(sleep_time):
        sleep_time2 = str(sleep_time.days)
        sleep_time3 = int(sleep_time2) * 86400  # 86,400 seconds in a day
        if sleep_time3 > 86400:
            message_body = f'There is currently no new data available. NJRScrapper will check again in {sleep_time.days} days...'
            Scraper.text_message(message_body)
            time.sleep(sleep_time3)
        else:
            message_body = f"There is currently no new data available. Will check again tomorrow..."
            Scraper.text_message(message_body)
            time.sleep(86400)

    @run_main
    @logger_decorator
    def main(self, **kwargs):
        global main_dictionary
        global full_year
        logger = kwargs['logger']
        f_handler = kwargs['f_handler']
        c_handler = kwargs['c_handler']
        # # If this code has never been run before, the full NJR10k will need to be run all the way back from 2018
        if self.no_of_runs == 0:

            first_results = self.njr10k()
            # The NJR10k function will return a list if the pdfs found to be possibly corrupted
            # If length of the list is created than 0,
            # the program will trigger the next function to download corrupted data
            if first_results == 'Read Logger File':
                # Read latest logger file to get a list of the corrupted files
                second_results = self.corrupted_files(self.read_logger(logger))
            elif first_results == 'All Files Downloaded':
                second_results = ['No Corrupted Files']
            elif len(first_results) > 0:
                second_results = self.corrupted_files(first_results)

            logger.info('Beginning PDF extraction...')
            time.sleep(1)
            for pdf in Scraper.pdf_generator():
                self.extract_re_data(pdf, second_results, main_dict=main_dictionary, year_dict=full_year)

            winsound.PlaySound('F:\\Python 2.0\\SoundFiles\\Victory.wav', 0)
            logger.info('PDF extraction is now complete...')

            old_dir = os.getcwd()
            # Use the Shelve module to save data for later use
            logger.info('Saving the data for Main Dictionary, Full Year and Event Log...')
            os.chdir('F:\\Python 2.0\\Projects\\Real Life Projects\\NJR Scrapper\\Saved Data')
            with shelve.open('NJ Scrapper Data Dictionary_v3') as saved_data_file:
                saved_data_file['Main Dictionary'] = main_dictionary
                saved_data_file['Full Year'] = full_year
                saved_data_file['Event Log'] = Scraper.event_log

        # If this code has been run before, the Updated NJR10k will need to be run from last pulled data
        elif self.no_of_runs > 0:

            start1, finish1 = Scraper.check_results()
            first_results = self.update_njr10k(start1, finish1)
            if first_results == 'Read Logger File':
                # Read latest logger file to get a list of the corrupted files
                second_results = self.corrupted_files(self.read_logger(logger))
            elif first_results == 'All Files Downloaded':
                second_results = ['No Corrupted Files']
            elif len(first_results) > 0:
                second_results = self.corrupted_files(first_results)

            old_dir = os.getcwd()
            # Use the Shelve module to save data for later use
            os.chdir('F:\\Python 2.0\\Projects\\Real Life Projects\\NJR Scrapper\\Saved Data')
            with shelve.open('NJ Scrapper Data Dictionary_v2') as saved_data_file:
                main_dictionary = saved_data_file['Main Dictionary']
                full_year = saved_data_file['Full Year']

            os.chdir(old_dir)
            logger.info('Beginning PDF extraction...')
            time.sleep(1)
            for pdf in Scraper.pdf_generator():
                self.extract_re_data(pdf, second_results, main_dict=main_dictionary, year_dict=full_year)

            winsound.PlaySound('F:\\Python 2.0\\SoundFiles\\Victory.wav', 0)
            logger.info('PDF extraction is now complete...')

            os.chdir('F:\\Python 2.0\\Projects\\Real Life Projects\\NJR Scrapper\\Saved Data')
            logger.info('Saving the data for Main Dictionary, Full Year and Event Log...')
            with shelve.open('NJ Scrapper Data Dictionary_v2', writeback=True) as saved_data_file:
                saved_data_file['Main Dictionary'] = main_dictionary
                saved_data_file['Full Year'] = full_year
                saved_data_file['Event Log'] = Scraper.event_log
                saved_data_file.sync()

        logger.info('All data has been saved...')
        os.chdir(old_dir)

        logger.info('Now sorting files into respective folders by year and township...')
        Scraper.OrganizeFiles(second_results)
        logger.info('All files have been sorted and organized...')

        now = datetime.datetime.strptime(time.ctime(), "%a %b %d %H:%M:%S %Y")
        end_of_year = datetime.datetime.strptime(Scraper.current_data.split(' ')[1] + "/12/31", "%Y/%m/%d")

        # If today's date is the last day of the year or greater run zip functino. If not, stay sleep
        if now >= end_of_year:
            logger.info('Now ziping files into respective folders by year and township')
            z_name = Scraper.CreateZip()
            logger.info(f'All files have been sorted and organized.\nFilename: {z_name}...')

        logger.info(f'Now converting Python dictionaries Main Dictionary and Full Year into Pandas dataframes\n'
                    f'Once complete, the dataframes will be transferred to an Excel file')
        excelfile = self.pandas2excel(main_dictionary, full_year)
        logger.info(f'All dataframes have been transferred to an Excel file\nFilename: {excelfile}')

        # Fix Scraper dictionary references
        # runs_list = [i for i in Scraper.event_log.keys()]
        message_body = '     NJ Scrapper     ' \
                       f"\n" \
                       f"\nEvent Log:" \
                       f"\nRun Date: {Scraper.event_log[Scraper.no_of_runs]['Run Date']}" \
                       f"\nRun #: {Scraper.no_of_runs}" \
                       f"\nRun Type: {Scraper.event_log[Scraper.no_of_runs]['Run Type']}" \
                       f"\nLatest Available Data: {Scraper.event_log[Scraper.no_of_runs]['Latest Available Data']}" \
                       f"\nRun Time: {Scraper.event_log[Scraper.no_of_runs]['Run Time']}" \
                       f"\nDays Between Update: {Scraper.event_log[Scraper.no_of_runs]['Days Between Update']}" \
                       f"\nFilename: {excelfile}"

        Scraper.text_message(message_body)
        logger.info('Program summary has been sent through text! The program is now complete!')
        logger.removeHandler(f_handler)
        logger.removeHandler(c_handler)
        logging.shutdown()

        return excelfile


if __name__ == '__main__':

    main_dictionary = {
        '2018': {},
        '2019': {},
        '2020': {},
        '2021': {},
        '2022': {},
        '2023': {}
    }

    full_year = {
        '2018': {},
        '2019': {},
        '2020': {},
        '2021': {},
        '2022': {},
        '2023': {}
    }

    """This is the first part of the NJ Realtor 10k Scrapper. 
    This section of the program will systematically check for the 
    most recent files uploaded to their database and download them 
    for processing and analysis in the second half of the program.
    """
    while True:
        try:
            obj = Scraper()
            # doc = 'F:\\Real Estate Investing\\JQH Holding Company LLC\\Real Estate Data\\NJ 10k Real Estate Data 2023-09-29.xlsx'
            # obj.cloropleth_maps_state(doc)
            # obj.matplot_lines(doc)
            results = obj.main()

            if results == 'RESTART':
                continue
            else:
                pass
            # run_quarterly_analysis(results)
        except KeyboardInterrupt:
            print()
            print('Program was manually stopped')






